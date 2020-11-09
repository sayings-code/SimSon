
# encoding:utf-8
import glob

import pandas as pd
import os
import xlrd
import openpyxl
from openpyxl.reader.excel import load_workbook
from sqlalchemy import create_engine



# 写入多个sheet 的方法：
# dataframe: 需要写入excel的数据
# outfile：输出的文件地址
# name: sheet_name的文件名称
def excelAddSheet(dataframe, outfile, name):
    writer = pd.ExcelWriter(outfile, enginge='openpyxl')
    if os.path.exists(outfile) != True:
        dataframe.to_excel(writer, name, index=None)
    else:
        book = load_workbook(writer.path)
        writer.book = book
        dataframe.to_excel(excel_writer=writer, sheet_name = name, index=None)
    writer.save()
    writer.close()

# 将多个报表整合成一个总表的过程
def totalReport(XLSX_NAME,SqlName):
    # 查找文件
    path_dir = 'D://develop//PycharmProjects//salereaport//collect//raw'
    files = glob.glob(os.path.join(path_dir, '*.xlsx'))
    # 整合的文件名
    FILE_NAME = 'collect/total.csv'

    # 读取【订单明细】工作表，第二行开始的内容，添加到列表
    SHEET_NAME = 'Sheet1'

    list1 = []

    for i in files:
        try:
            print(f'正在处理: {i}')
            m = pd.read_excel(i, header=1, sheet_name="订单明细")
            df = m[
                ["订单时间", "发货时间", "店铺", "订单号", "产品", "产品2", "数量", "零售单价", "币种1", "佣金比例", "结算总额", "币种2", "零售单价¥", "结算金额￥",
                 "成本价", "币种3", "成本总价￥", "蚂蚁操作费", "头/全程运费￥", "尾程运费￥", "资金成本", "毛利￥", "发货仓位", "收件人国家", "物流单号", "回款情况"]]
            # print(i,set([type(i) for i in m['订单时间'].tolist()]))
            # print(i,m.columns)
            list1.append(df)
        except Exception as e:
            print(f'{i}文件出错了！！')
            print(e)
            pass

    # 开始拼接
    zong = pd.concat(list1, sort=True)

    # 按顺序（order列表顺序）写入csv文件
    order = ["订单时间", "发货时间", "店铺", "订单号", "产品", "产品2", "数量", "零售单价", "币种1", "佣金比例", "结算总额", "币种2", "零售单价¥", "结算金额￥",
             "成本价", "币种3", "成本总价￥", "蚂蚁操作费", "头/全程运费￥", "尾程运费￥", "资金成本", "毛利￥", "发货仓位", "收件人国家", "物流单号", "回款情况"]
    zong = zong[order]
    # print([i for i in zong['订单时间'].tolist()])
    zong.sort_values(by='订单时间', inplace=True)
    # print(zong)
    zong.to_csv(FILE_NAME, index=False, encoding='utf_8_sig')

    # wb = openpyxl.load_workbook(XLSX_NAME)
    # wb.remove(wb[SHEET_NAME])
    # wb.create_sheet(index=0, title=SHEET_NAME)
    # sheet = wb[SHEET_NAME]

    titles = (
        "订单时间", "发货时间", "店铺", "订单号", "产品", "产品2", "数量", "零售单价", "币种1", "佣金比例", "结算总额", "币种2", "零售单价¥", "结算金额￥", "成本价",
        "币种3", "成本总价￥", "蚂蚁操作费", "头/全程运费￥", "尾程运费￥", "资金成本", "毛利￥", "发货仓位", "收件人国家", "物流单号", "回款情况")

    # 读取total.csv
    data = pd.read_csv(FILE_NAME, header=0)

    # List = data.values.tolist()
    # # 先写入表头
    # sheet.append(titles)
    #
    # for i in List:
    #     i = tuple(i)
    #     sheet.append(i)
    #
    # wb.save(XLSX_NAME)

    print("over，over！！")

    data['店铺'] = data['店铺'].str.lower()  # 店铺小写

    # 整合部门表
    DEPART = 'D://develop//PycharmProjects//salereaport//collect//11月部门架构.xlsx'
    df_depart = pd.read_excel(DEPART, header=0)

    df_depart['Store'] = df_depart['Store'].str.lower()  # 店铺小写
    # 根据店铺整合
    result = pd.merge(data, df_depart, how='left', left_on='店铺', right_on='Store')

    # 总数据存入数据库
    engine = create_engine('mysql+pymysql://root:root@localhost:3306/test')
    result.drop('updatetime', inplace=True, axis=1)
    result.drop('Store', inplace=True, axis=1)
    result.rename(columns={'Operator': '负责人', 'Platform': '平台'}, inplace=True)
    result.to_sql(SqlName, con=engine, schema='test', if_exists='replace', index=0)
    print("--当日总数据入库成功！--")

    # 写入excel
    result.to_excel(XLSX_NAME, index=0)
    print("==成功生成当日总数据表！==")


# 部门表入库，一般一月使用一次
def departToSql(DEPART,TABLENAME):
    df_depart = pd.read_excel(DEPART, header=0)
    # 1.部门表入库
    engine = create_engine('mysql+pymysql://root:root@localhost:3306/test')
    df_depart.to_sql(TABLENAME, con=engine, schema='test', if_exists='replace', index=0)
    print("--部门表入库成功！--")

# 两张数据表联合操作： (常用！)
def concatTwoDf(df1,df2,Output_Data,SqlName):
    # 整合两个表的数据：
    Data1 = pd.read_excel(df1, header=0)
    Data2 = pd.read_excel(df2, header=0)

    Data_total = pd.concat([Data2, Data1])

    # 写入excel
    Data_total.to_excel(Output_Data, index=0)

    #添加部门sheet
    df_depart = pd.read_excel('D:/develop/PycharmProjects/salereaport/collect/11月部门架构.xlsx', header=0)
    df_depart.drop('updatetime',axis=1,inplace=True)
    excelAddSheet(df_depart,Output_Data,'部门')

    print('---整合数据表成功输出！---')

    # 总数据存入数据库
    engine = create_engine('mysql+pymysql://root:root@localhost:3306/test')
    Data1.to_sql(SqlName, con=engine, schema='test', if_exists='replace', index=0)
    print("--当日总数据入库成功！--")


# 改造不完整的表 -> 总数据表，一般不用
def RepariTable(TABLEPATH,DEPARTPATH,OUTPATH):
    #改造10月份表的代码
    data = pd.read_excel(TABLEPATH, header=0)
    DEPART = DEPARTPATH
    df_depart = pd.read_excel(DEPART, header=0)
    df_depart['store'] = df_depart['store'].str.lower()  # 店铺小写
    data['店铺'] = data['店铺'].str.lower()  # 店铺小写
    # 根据店铺整合
    result = pd.merge(data, df_depart, how='left', left_on='店铺', right_on='store')
    result.drop('updatetime', inplace=True, axis=1)
    result.drop('store', inplace=True, axis=1)
    result.rename(columns={'operator': '运营负责人', 'depart': '三级部门', 'platform': '平台'}, inplace=True)
    result.to_excel(OUTPATH,index=0)
    print("--原有表改造完成！--")